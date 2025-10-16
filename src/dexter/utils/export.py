import io
from datetime import datetime
from typing import Dict, List, Any
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExportManager:
    """
    Export research results to PDF and Excel formats.
    """
    
    @staticmethod
    def generate_pdf(query: str, answer: str, tasks: List[Dict], stats: Dict[str, Any]) -> bytes:
        """Generate a PDF report of research results."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)
        
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#2ca02c'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        story.append(Paragraph("Dexter Financial Research Report", title_style))
        story.append(Spacer(1, 12))
        
        # Timestamp
        timestamp = datetime.now().strftime("%B %d, %Y at %H:%M")
        story.append(Paragraph(f"Generated: {timestamp}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Query section
        story.append(Paragraph("Research Query", heading_style))
        story.append(Paragraph(query, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary section
        story.append(Paragraph("Executive Summary", heading_style))
        # Split answer into paragraphs for better formatting
        for para in answer.split('\n\n'):
            if para.strip():
                story.append(Paragraph(para.strip(), styles['Normal']))
                story.append(Spacer(1, 6))
        story.append(Spacer(1, 20))
        
        # Statistics section
        if stats:
            story.append(Paragraph("Execution Statistics", heading_style))
            stats_data = [
                ['Metric', 'Value'],
                ['Total Steps', str(stats.get('total_steps', 'N/A'))],
                ['Tasks Completed', str(stats.get('tasks_completed', 'N/A'))],
                ['API Calls', str(stats.get('api_calls', 'N/A'))],
                ['Execution Time', f"{stats.get('execution_time', 0):.2f}s"]
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 20))
        
        # Tasks breakdown
        if tasks:
            story.append(PageBreak())
            story.append(Paragraph("Detailed Research Breakdown", heading_style))
            story.append(Spacer(1, 12))
            
            for i, task in enumerate(tasks, 1):
                story.append(Paragraph(f"Task {i}: {task.get('description', 'Unknown')}", 
                                      styles['Heading3']))
                story.append(Spacer(1, 6))
                
                if task.get('result'):
                    result_text = task['result'][:500] + "..." if len(task['result']) > 500 else task['result']
                    story.append(Paragraph(result_text, styles['Normal']))
                    story.append(Spacer(1, 12))
        
        # Build PDF
        doc.build(story)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    @staticmethod
    def generate_excel(query: str, answer: str, tasks: List[Dict], stats: Dict[str, Any]) -> bytes:
        """Generate an Excel workbook of research results."""
        workbook = Workbook()
        
        # Summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        title_font = Font(bold=True, size=16)
        
        # Title
        summary_sheet['A1'] = "Dexter Financial Research Report"
        summary_sheet['A1'].font = title_font
        summary_sheet.merge_cells('A1:D1')
        
        # Timestamp
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
        summary_sheet['A3'] = ""
        
        # Query
        summary_sheet['A4'] = "Research Query"
        summary_sheet['A4'].font = header_font
        summary_sheet['A4'].fill = header_fill
        summary_sheet['A5'] = query
        summary_sheet.merge_cells('A5:D5')
        summary_sheet['A5'].alignment = Alignment(wrap_text=True, vertical='top')
        summary_sheet['A6'] = ""
        
        # Summary
        summary_sheet['A7'] = "Executive Summary"
        summary_sheet['A7'].font = header_font
        summary_sheet['A7'].fill = header_fill
        summary_sheet['A8'] = answer[:32000]  # Excel cell limit
        summary_sheet.merge_cells('A8:D8')
        summary_sheet['A8'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 25
        summary_sheet.column_dimensions['C'].width = 25
        summary_sheet.column_dimensions['D'].width = 25
        
        # Statistics sheet
        if stats:
            stats_sheet = workbook.create_sheet("Statistics")
            stats_sheet['A1'] = "Execution Statistics"
            stats_sheet['A1'].font = title_font
            stats_sheet['A2'] = ""
            
            stats_sheet['A3'] = "Metric"
            stats_sheet['B3'] = "Value"
            stats_sheet['A3'].font = header_font
            stats_sheet['B3'].font = header_font
            stats_sheet['A3'].fill = header_fill
            stats_sheet['B3'].fill = header_fill
            
            row = 4
            for key, value in stats.items():
                stats_sheet[f'A{row}'] = key.replace('_', ' ').title()
                stats_sheet[f'B{row}'] = str(value) if not isinstance(value, float) else f"{value:.2f}"
                row += 1
            
            stats_sheet.column_dimensions['A'].width = 20
            stats_sheet.column_dimensions['B'].width = 20
        
        # Tasks sheet
        if tasks:
            tasks_sheet = workbook.create_sheet("Research Tasks")
            tasks_sheet['A1'] = "Research Task Breakdown"
            tasks_sheet['A1'].font = title_font
            tasks_sheet['A2'] = ""
            
            tasks_sheet['A3'] = "Task #"
            tasks_sheet['B3'] = "Description"
            tasks_sheet['C3'] = "Result"
            for col in ['A3', 'B3', 'C3']:
                tasks_sheet[col].font = header_font
                tasks_sheet[col].fill = header_fill
            
            row = 4
            for i, task in enumerate(tasks, 1):
                tasks_sheet[f'A{row}'] = i
                tasks_sheet[f'B{row}'] = task.get('description', 'Unknown')
                tasks_sheet[f'C{row}'] = task.get('result', '')[:32000]
                tasks_sheet[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                tasks_sheet[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
            
            tasks_sheet.column_dimensions['A'].width = 10
            tasks_sheet.column_dimensions['B'].width = 40
            tasks_sheet.column_dimensions['C'].width = 50
        
        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes
